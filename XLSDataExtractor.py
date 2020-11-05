from openpyxl import load_workbook
import pandas as pd


# wb = load_workbook(filename=r'D:\Work\Scripts\Data\XLSDataExtractor\Strategy_for_TOPIX.xlsx')
# sheet = wb['data']
#
# dates = sheet['A2:A3510']
# values = sheet['D2:D3510']
#
# result = []
#
# for date, value in zip(dates, values):
#     date = pd.to_datetime(date[0].value)
#     result.append((date, value[0].value))
#
# df = pd.DataFrame(result, columns=['Date', 'value'])
# df = df.set_index('Date')
# df.to_csv('test.csv')


def regime_export():
    wb = load_workbook(filename=r'C:\Users\Admin\Downloads\20201016102133\Strategy_for_TOPIX.xlsx')
    sheet = wb['data']

    mapping = {
        1.5: 4,
        1: 3,
        0.5: 2,
        0: 1,
        -0.5: 0
    }

    dates = sheet['A2:A3545']
    values = sheet['B2:B3545']

    result = []

    for date, value in zip(dates, values):
        date = pd.to_datetime(date[0].value)
        result.append((date, mapping[value[0].value]))

    df = pd.DataFrame(result, columns=['Date', 'value'])
    df = df.set_index('Date')

    for num in range(len(mapping.values())):
        df[str(num)] = df.apply(lambda row: row['value'] == num, axis=1)
    df = df.drop(['value'], axis=1)

    df = df.astype(int)

    df.to_csv('test.csv')


regime_export()
